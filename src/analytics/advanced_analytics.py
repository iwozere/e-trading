"""
Advanced Analytics & Reporting System
====================================

Provides comprehensive analytics including:
- Advanced performance metrics
- Monte Carlo simulations
- Risk analysis (VaR, CVaR)
- Strategy comparison and ranking
- Automated reporting with PDF/Excel export
- Portfolio analytics and correlation analysis
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple, Any
import json
import os
import warnings
warnings.filterwarnings('ignore')

from src.model.analytics import PerformanceMetrics, Trade

# For PDF generation
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# For Excel export
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False




class AdvancedAnalytics:
    """
    Advanced analytics system for trading strategy analysis
    """

    def __init__(self, risk_free_rate: float = 0.02):
        """
        Initialize the analytics system

        Args:
            risk_free_rate: Annual risk-free rate (default: 2%)
        """
        self.risk_free_rate = risk_free_rate
        self.trades: List[Trade] = []
        self.metrics: Optional[PerformanceMetrics] = None

    def add_trades(self, trades_data: List[Dict[str, Any]]):
        """
        Add trades from optimization results or database

        Args:
            trades_data: List of trade dictionaries
        """
        for trade_data in trades_data:
            trade = Trade(
                entry_time=datetime.fromisoformat(trade_data['entry_time'].replace('Z', '+00:00')),
                exit_time=datetime.fromisoformat(trade_data['exit_time'].replace('Z', '+00:00')),
                symbol=trade_data.get('symbol', 'Unknown'),
                side=trade_data.get('side', 'Unknown'),
                entry_price=float(trade_data['entry_price']),
                exit_price=float(trade_data['exit_price']),
                quantity=float(trade_data['quantity']),
                pnl=float(trade_data.get('pnl', 0)),
                commission=float(trade_data.get('commission', 0)),
                net_pnl=float(trade_data['net_pnl']),
                exit_reason=trade_data.get('exit_reason', 'unknown')
            )
            self.trades.append(trade)

    def calculate_metrics(self) -> PerformanceMetrics:
        """
        Calculate comprehensive performance metrics

        Returns:
            PerformanceMetrics object with all calculated metrics
        """
        if not self.trades:
            return PerformanceMetrics()

        # Sort trades by exit time
        self.trades.sort(key=lambda x: x.exit_time)

        # Basic trade analysis
        total_trades = len(self.trades)
        winning_trades = [t for t in self.trades if t.net_pnl > 0]
        losing_trades = [t for t in self.trades if t.net_pnl <= 0]

        win_rate = len(winning_trades) / total_trades * 100 if total_trades > 0 else 0

        # Profit/Loss analysis
        total_return = sum(t.net_pnl for t in self.trades)
        gross_profit = sum(t.net_pnl for t in winning_trades)
        gross_loss = abs(sum(t.net_pnl for t in losing_trades))

        profit_factor = gross_profit / gross_loss if gross_loss > 0 else float('inf')

        # Calculate portfolio value series
        portfolio_values = self._calculate_portfolio_values()

        # Risk metrics
        max_drawdown, max_drawdown_pct = self._calculate_max_drawdown(portfolio_values)
        sharpe_ratio = self._calculate_sharpe_ratio(portfolio_values)
        sortino_ratio = self._calculate_sortino_ratio(portfolio_values)
        calmar_ratio = self._calculate_calmar_ratio(portfolio_values, max_drawdown_pct)

        # VaR and CVaR
        var_95, cvar_95 = self._calculate_var_cvar()

        # Kelly Criterion
        kelly_criterion = self._calculate_kelly_criterion()

        # Trade analysis
        avg_win = np.mean([t.net_pnl for t in winning_trades]) if winning_trades else 0
        avg_loss = np.mean([t.net_pnl for t in losing_trades]) if losing_trades else 0
        largest_win = max([t.net_pnl for t in self.trades]) if self.trades else 0
        largest_loss = min([t.net_pnl for t in self.trades]) if self.trades else 0

        # Duration analysis
        durations = [t.duration for t in self.trades]
        avg_trade_duration = np.mean(durations) if durations else timedelta()

        # Consecutive analysis
        max_consecutive_wins, max_consecutive_losses = self._calculate_consecutive_trades()

        # Time analysis
        trading_days = self._calculate_trading_days()
        avg_trades_per_day = total_trades / trading_days if trading_days > 0 else 0

        # Additional metrics
        recovery_factor = total_return / abs(max_drawdown) if max_drawdown != 0 else 0
        payoff_ratio = avg_win / abs(avg_loss) if avg_loss != 0 else 0
        expectancy = (win_rate / 100 * avg_win) + ((1 - win_rate / 100) * avg_loss)

        self.metrics = PerformanceMetrics(
            total_trades=total_trades,
            winning_trades=len(winning_trades),
            losing_trades=len(losing_trades),
            win_rate=win_rate,
            profit_factor=profit_factor,
            total_return=total_return,
            total_return_pct=total_return / 1000 * 100,  # Assuming 1000 initial capital
            max_drawdown=max_drawdown,
            max_drawdown_pct=max_drawdown_pct,
            sharpe_ratio=sharpe_ratio,
            sortino_ratio=sortino_ratio,
            calmar_ratio=calmar_ratio,
            var_95=var_95,
            cvar_95=cvar_95,
            kelly_criterion=kelly_criterion,
            expectancy=expectancy,
            avg_win=avg_win,
            avg_loss=avg_loss,
            largest_win=largest_win,
            largest_loss=largest_loss,
            avg_trade_duration=avg_trade_duration,
            max_consecutive_wins=max_consecutive_wins,
            max_consecutive_losses=max_consecutive_losses,
            avg_trades_per_day=avg_trades_per_day,
            recovery_factor=recovery_factor,
            payoff_ratio=payoff_ratio,
            profit_factor_ratio=profit_factor
        )

        return self.metrics

    def _calculate_portfolio_values(self) -> List[float]:
        """Calculate portfolio value series"""
        if not self.trades:
            return [1000.0]  # Default initial value

        initial_capital = 1000.0
        portfolio_values = [initial_capital]
        current_value = initial_capital

        # Group trades by date
        daily_pnl = {}
        for trade in self.trades:
            trade_date = trade.exit_time.date()
            if trade_date not in daily_pnl:
                daily_pnl[trade_date] = 0.0
            daily_pnl[trade_date] += trade.net_pnl

        # Create portfolio value series
        dates = sorted(daily_pnl.keys())
        for date in dates:
            current_value += daily_pnl[date]
            portfolio_values.append(current_value)

        return portfolio_values

    def _calculate_max_drawdown(self, portfolio_values: List[float]) -> Tuple[float, float]:
        """Calculate maximum drawdown"""
        if len(portfolio_values) < 2:
            return 0.0, 0.0

        peak = portfolio_values[0]
        max_dd = 0.0
        max_dd_pct = 0.0

        for value in portfolio_values:
            if value > peak:
                peak = value
            drawdown = peak - value
            drawdown_pct = (drawdown / peak) * 100

            if drawdown > max_dd:
                max_dd = drawdown
                max_dd_pct = drawdown_pct

        return max_dd, max_dd_pct

    def _calculate_sharpe_ratio(self, portfolio_values: List[float]) -> float:
        """Calculate Sharpe ratio"""
        if len(portfolio_values) < 2:
            return 0.0

        returns = []
        for i in range(1, len(portfolio_values)):
            if portfolio_values[i-1] != 0:
                ret = (portfolio_values[i] - portfolio_values[i-1]) / portfolio_values[i-1]
                returns.append(ret)

        if not returns:
            return 0.0

        returns = np.array(returns)
        avg_return = np.mean(returns)
        std_return = np.std(returns, ddof=1)

        # Annualize
        annualized_return = avg_return * 252
        annualized_volatility = std_return * np.sqrt(252)

        if annualized_volatility == 0:
            return 0.0

        return (annualized_return - self.risk_free_rate) / annualized_volatility

    def _calculate_sortino_ratio(self, portfolio_values: List[float]) -> float:
        """Calculate Sortino ratio"""
        if len(portfolio_values) < 2:
            return 0.0

        returns = []
        for i in range(1, len(portfolio_values)):
            if portfolio_values[i-1] != 0:
                ret = (portfolio_values[i] - portfolio_values[i-1]) / portfolio_values[i-1]
                returns.append(ret)

        if not returns:
            return 0.0

        returns = np.array(returns)
        avg_return = np.mean(returns)

        # Calculate downside deviation
        downside_returns = returns[returns < self.risk_free_rate/252]
        if len(downside_returns) == 0:
            return float('inf')

        downside_dev = np.sqrt(np.mean((downside_returns - self.risk_free_rate/252) ** 2))

        if downside_dev == 0:
            return float('inf')

        # Annualize
        annualized_return = avg_return * 252
        annualized_downside_dev = downside_dev * np.sqrt(252)

        return (annualized_return - self.risk_free_rate) / annualized_downside_dev

    def _calculate_calmar_ratio(self, portfolio_values: List[float], max_dd_pct: float) -> float:
        """Calculate Calmar ratio"""
        if len(portfolio_values) < 2 or max_dd_pct == 0:
            return 0.0

        total_return_pct = ((portfolio_values[-1] - portfolio_values[0]) / portfolio_values[0]) * 100
        return (total_return_pct - self.risk_free_rate * 100) / max_dd_pct

    def _calculate_var_cvar(self) -> Tuple[float, float]:
        """Calculate Value at Risk and Conditional Value at Risk"""
        if not self.trades:
            return 0.0, 0.0

        returns = [t.net_pnl for t in self.trades]
        returns = np.array(returns)

        var_95 = np.percentile(returns, 5)  # 95% VaR
        cvar_95 = np.mean(returns[returns <= var_95])  # 95% CVaR

        return var_95, cvar_95

    def _calculate_kelly_criterion(self) -> float:
        """Calculate Kelly Criterion"""
        if not self.trades:
            return 0.0

        winning_trades = [t for t in self.trades if t.net_pnl > 0]
        losing_trades = [t for t in self.trades if t.net_pnl <= 0]

        if not winning_trades or not losing_trades:
            return 0.0

        win_rate = len(winning_trades) / len(self.trades)
        avg_win = np.mean([t.net_pnl for t in winning_trades])
        avg_loss = abs(np.mean([t.net_pnl for t in losing_trades]))

        if avg_loss == 0:
            return 0.0

        kelly = (win_rate * avg_win - (1 - win_rate) * avg_loss) / avg_win
        return max(0, min(kelly, 1))  # Clamp between 0 and 1

    def _calculate_consecutive_trades(self) -> Tuple[int, int]:
        """Calculate maximum consecutive wins and losses"""
        if not self.trades:
            return 0, 0

        current_wins = 0
        current_losses = 0
        max_wins = 0
        max_losses = 0

        for trade in self.trades:
            if trade.net_pnl > 0:
                current_wins += 1
                current_losses = 0
                max_wins = max(max_wins, current_wins)
            else:
                current_losses += 1
                current_wins = 0
                max_losses = max(max_losses, current_losses)

        return max_wins, max_losses

    def _calculate_trading_days(self) -> int:
        """Calculate number of trading days"""
        if not self.trades:
            return 0

        start_date = min(t.entry_time.date() for t in self.trades)
        end_date = max(t.exit_time.date() for t in self.trades)

        return (end_date - start_date).days + 1

    def run_monte_carlo_simulation(self, n_simulations: int = 10000, n_trades: int = 100) -> Dict[str, Any]:
        """
        Run Monte Carlo simulation to estimate future performance

        Args:
            n_simulations: Number of simulations to run
            n_trades: Number of trades per simulation

        Returns:
            Dictionary with simulation results
        """
        if not self.trades:
            return {"error": "No trades available for simulation"}

        # Extract trade returns
        trade_returns = [t.net_pnl for t in self.trades]

        if len(trade_returns) < 10:
            return {"error": "Insufficient trade data for simulation"}

        # Calculate empirical distribution parameters
        mean_return = np.mean(trade_returns)
        std_return = np.std(trade_returns, ddof=1)

        # Run simulations
        simulation_results = []

        for _ in range(n_simulations):
            # Generate random trade returns
            simulated_returns = np.random.normal(mean_return, std_return, n_trades)
            cumulative_return = np.sum(simulated_returns)
            simulation_results.append(cumulative_return)

        simulation_results = np.array(simulation_results)

        # Calculate statistics
        mean_sim_return = np.mean(simulation_results)
        std_sim_return = np.std(simulation_results)
        var_95_sim = np.percentile(simulation_results, 5)
        cvar_95_sim = np.mean(simulation_results[simulation_results <= var_95_sim])

        # Calculate probability of profit
        prob_profit = np.mean(simulation_results > 0) * 100

        return {
            "mean_return": mean_sim_return,
            "std_return": std_sim_return,
            "var_95": var_95_sim,
            "cvar_95": cvar_95_sim,
            "prob_profit": prob_profit,
            "min_return": np.min(simulation_results),
            "max_return": np.max(simulation_results),
            "percentiles": {
                "10": np.percentile(simulation_results, 10),
                "25": np.percentile(simulation_results, 25),
                "50": np.percentile(simulation_results, 50),
                "75": np.percentile(simulation_results, 75),
                "90": np.percentile(simulation_results, 90)
            }
        }

    def generate_performance_report(self, output_dir: str = "reports") -> str:
        """
        Generate comprehensive performance report

        Args:
            output_dir: Directory to save reports

        Returns:
            Path to generated report
        """
        if not self.metrics:
            self.calculate_metrics()

        # Create output directory
        os.makedirs(output_dir, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        # Generate different report formats
        reports = []

        if REPORTLAB_AVAILABLE:
            pdf_path = os.path.join(output_dir, f"performance_report_{timestamp}.pdf")
            self._generate_pdf_report(pdf_path)
            reports.append(pdf_path)

        if OPENPYXL_AVAILABLE:
            excel_path = os.path.join(output_dir, f"performance_report_{timestamp}.xlsx")
            self._generate_excel_report(excel_path)
            reports.append(excel_path)

        # Always generate JSON report
        json_path = os.path.join(output_dir, f"performance_report_{timestamp}.json")
        self._generate_json_report(json_path)
        reports.append(json_path)

        return reports[0] if reports else json_path

    def _generate_pdf_report(self, filepath: str):
        """Generate PDF performance report"""
        if not REPORTLAB_AVAILABLE:
            return

        doc = SimpleDocTemplate(filepath, pagesize=A4)
        story = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1  # Center
        )
        story.append(Paragraph("Trading Strategy Performance Report", title_style))
        story.append(Spacer(1, 12))

        # Summary metrics
        story.append(Paragraph("Performance Summary", styles['Heading2']))
        story.append(Spacer(1, 12))

        summary_data = [
            ['Metric', 'Value'],
            ['Total Trades', f"{self.metrics.total_trades}"],
            ['Win Rate', f"{self.metrics.win_rate:.2f}%"],
            ['Profit Factor', f"{self.metrics.profit_factor:.2f}"],
            ['Total Return', f"${self.metrics.total_return:.2f}"],
            ['Sharpe Ratio', f"{self.metrics.sharpe_ratio:.2f}"],
            ['Max Drawdown', f"{self.metrics.max_drawdown_pct:.2f}%"],
            ['Calmar Ratio', f"{self.metrics.calmar_ratio:.2f}"],
        ]

        summary_table = Table(summary_data)
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(summary_table)

        doc.build(story)

    def _generate_excel_report(self, filepath: str):
        """Generate Excel performance report"""
        if not OPENPYXL_AVAILABLE:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Performance Summary"

        # Headers
        headers = ['Metric', 'Value', 'Target', 'Status']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
            ws.cell(row=1, column=col).font = Font(bold=True)
            ws.cell(row=1, column=col).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Data
        data = [
            ['Total Trades', self.metrics.total_trades, '>30', '✓' if self.metrics.total_trades > 30 else '✗'],
            ['Win Rate (%)', round(self.metrics.win_rate, 2), '>50', '✓' if self.metrics.win_rate > 50 else '✗'],
            ['Profit Factor', round(self.metrics.profit_factor, 2), '>1.5', '✓' if self.metrics.profit_factor > 1.5 else '✗'],
            ['Sharpe Ratio', round(self.metrics.sharpe_ratio, 2), '>1.0', '✓' if self.metrics.sharpe_ratio > 1.0 else '✗'],
            ['Max Drawdown (%)', round(self.metrics.max_drawdown_pct, 2), '<20', '✓' if self.metrics.max_drawdown_pct < 20 else '✗'],
            ['Calmar Ratio', round(self.metrics.calmar_ratio, 2), '>1.0', '✓' if self.metrics.calmar_ratio > 1.0 else '✗'],
            ['Sortino Ratio', round(self.metrics.sortino_ratio, 2), '>1.0', '✓' if self.metrics.sortino_ratio > 1.0 else '✗'],
            ['Recovery Factor', round(self.metrics.recovery_factor, 2), '>1.0', '✓' if self.metrics.recovery_factor > 1.0 else '✗'],
        ]

        for row, (metric, value, target, status) in enumerate(data, 2):
            ws.cell(row=row, column=1, value=metric)
            ws.cell(row=row, column=2, value=value)
            ws.cell(row=row, column=3, value=target)
            ws.cell(row=row, column=4, value=status)

            # Color code status
            if status == '✓':
                ws.cell(row=row, column=4).fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            else:
                ws.cell(row=row, column=4).fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(filepath)

    def _generate_json_report(self, filepath: str):
        """Generate JSON performance report"""
        report_data = {
            "report_generated": datetime.now().isoformat(),
            "strategy_info": {
                "total_trades": self.metrics.total_trades,
                "analysis_period": {
                    "start": min(t.entry_time for t in self.trades).isoformat() if self.trades else None,
                    "end": max(t.exit_time for t in self.trades).isoformat() if self.trades else None
                }
            },
            "performance_metrics": {
                "basic_metrics": {
                    "win_rate": round(self.metrics.win_rate, 2),
                    "profit_factor": round(self.metrics.profit_factor, 2),
                    "total_return": round(self.metrics.total_return, 2),
                    "total_return_pct": round(self.metrics.total_return_pct, 2)
                },
                "risk_metrics": {
                    "max_drawdown": round(self.metrics.max_drawdown, 2),
                    "max_drawdown_pct": round(self.metrics.max_drawdown_pct, 2),
                    "sharpe_ratio": round(self.metrics.sharpe_ratio, 2),
                    "sortino_ratio": round(self.metrics.sortino_ratio, 2),
                    "calmar_ratio": round(self.metrics.calmar_ratio, 2),
                    "var_95": round(self.metrics.var_95, 2),
                    "cvar_95": round(self.metrics.cvar_95, 2)
                },
                "trade_analysis": {
                    "avg_win": round(self.metrics.avg_win, 2),
                    "avg_loss": round(self.metrics.avg_loss, 2),
                    "largest_win": round(self.metrics.largest_win, 2),
                    "largest_loss": round(self.metrics.largest_loss, 2),
                    "max_consecutive_wins": self.metrics.max_consecutive_wins,
                    "max_consecutive_losses": self.metrics.max_consecutive_losses
                }
            },
            "recommendations": self._generate_recommendations()
        }

        with open(filepath, 'w') as f:
            json.dump(report_data, f, indent=2, default=str)

    def _generate_recommendations(self) -> List[str]:
        """Generate trading recommendations based on metrics"""
        recommendations = []

        if self.metrics.win_rate < 40:
            recommendations.append("Consider improving entry criteria - win rate is below 40%")
        elif self.metrics.win_rate > 70:
            recommendations.append("Excellent win rate! Consider increasing position sizes")

        if self.metrics.profit_factor < 1.2:
            recommendations.append("Profit factor is low - review exit strategies and risk management")
        elif self.metrics.profit_factor > 2.0:
            recommendations.append("Strong profit factor - strategy is performing well")

        if self.metrics.max_drawdown_pct > 20:
            recommendations.append("High drawdown detected - implement stricter risk management")

        if self.metrics.sharpe_ratio < 1.0:
            recommendations.append("Sharpe ratio below 1.0 - consider risk-adjusted improvements")

        if self.metrics.total_trades < 30:
            recommendations.append("Limited trade sample - need more data for reliable analysis")

        if not recommendations:
            recommendations.append("Strategy is performing well across all metrics")

        return recommendations


class StrategyComparator:
    """
    Compare multiple trading strategies
    """

    def __init__(self):
        self.strategies: Dict[str, AdvancedAnalytics] = {}

    def add_strategy(self, name: str, analytics: AdvancedAnalytics):
        """Add a strategy for comparison"""
        self.strategies[name] = analytics

    def compare_strategies(self) -> pd.DataFrame:
        """Compare all strategies and return comparison table"""
        if not self.strategies:
            return pd.DataFrame()

        comparison_data = []

        for name, analytics in self.strategies.items():
            if not analytics.metrics:
                analytics.calculate_metrics()

            metrics = analytics.metrics
            comparison_data.append({
                'Strategy': name,
                'Total Trades': metrics.total_trades,
                'Win Rate (%)': round(metrics.win_rate, 2),
                'Profit Factor': round(metrics.profit_factor, 2),
                'Total Return ($)': round(metrics.total_return, 2),
                'Sharpe Ratio': round(metrics.sharpe_ratio, 2),
                'Max Drawdown (%)': round(metrics.max_drawdown_pct, 2),
                'Calmar Ratio': round(metrics.calmar_ratio, 2),
                'Sortino Ratio': round(metrics.sortino_ratio, 2),
                'Recovery Factor': round(metrics.recovery_factor, 2)
            })

        df = pd.DataFrame(comparison_data)
        return df.sort_values('Sharpe Ratio', ascending=False)

    def rank_strategies(self) -> Dict[str, int]:
        """Rank strategies by overall performance score"""
        if not self.strategies:
            return {}

        scores = {}

        for name, analytics in self.strategies.items():
            if not analytics.metrics:
                analytics.calculate_metrics()

            metrics = analytics.metrics

            # Calculate composite score (weighted average)
            score = (
                metrics.win_rate * 0.2 +
                min(metrics.profit_factor, 5.0) * 20 * 0.2 +
                metrics.sharpe_ratio * 10 * 0.2 +
                (100 - metrics.max_drawdown_pct) * 0.2 +
                metrics.calmar_ratio * 10 * 0.2
            )

            scores[name] = score

        # Sort by score
        sorted_scores = sorted(scores.items(), key=lambda x: x[1], reverse=True)

        # Return rankings
        rankings = {}
        for i, (name, score) in enumerate(sorted_scores, 1):
            rankings[name] = i

        return rankings
